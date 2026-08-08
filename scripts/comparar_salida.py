#!/usr/bin/env python
"""Compara la salida del script contra el Excel de referencia hecho a mano.

Uso:
    .venv/bin/python scripts/comparar_salida.py [salida.xlsx] [referencia.xlsx]

La referencia es ANCHA (un expediente por fila, MOTIVO 1 / MOTIVO 2 en columnas
separadas). La salida es LARGA (una fila por causal). Se normaliza agrupando
nuestras filas por expediente y comparando CONJUNTOS de motivos.
"""
import re
import sys
import unicodedata
from collections import defaultdict

import openpyxl

REF_DEFAULT = "Negacion marcas con información extra.xlsx"
OUT_DEFAULT = "Negacion_marcas_20260807_213955.xlsx"

# columnas 1-based
REF = dict(exp=1, marca=2, naturaleza=3, oposicion=4, op1=5, corto1=6, art1=7,
           fund1=8, op2=9, corto2=10, art2=11, fund2=12, mot1=13, mot2=14,
           apelacion=15, titular=16, niza=17, desc=18)
OUT = dict(exp=1, marca=2, naturaleza=3, oposicion=4, op1=5, corto1=6, art1=7,
           fund1=8, op2=9, corto2=10, art2=11, fund2=12, motivo=13,
           apelacion=14, titular=15, niza=16, mot_n=17, mot_tot=18, obs=19,
           desc=20)

HYPERLINK = re.compile(r'=HYPERLINK\([^,]*,\s*"([^"]*)"\s*\)', re.I)


def sin_tildes(s):
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def norm(v):
    """Mayusculas, sin tildes, espacios colapsados. '' si vacio."""
    if v is None:
        return ""
    s = sin_tildes(str(v)).upper()
    s = re.sub(r"\s+", " ", s).strip(" .,;")
    return s


def norm_si_no(v):
    s = norm(v)
    if s in ("SI", "S", "X", "TRUE", "VERDADERO"):
        return "SI"
    if s in ("NO", "N", "FALSE", "FALSO"):
        return "NO"
    return s


def norm_exp(v):
    """SD2022/0000017 -> SD2022-0000017 (clave canonica)."""
    s = norm(v).replace("/", "-").replace(" ", "")
    return s


def norm_motivo(v):
    """136a / 136 A / Art. 136 literal a -> 136A."""
    s = norm(v).replace(" ", "").replace(".", "").replace("-", "")
    s = s.replace("ARTICULO", "").replace("ART", "").replace("LITERAL", "")
    return s


def leer_ref(path):
    """La referencia tambien es parcialmente LARGA: hay filas sin numero de
    expediente que son continuacion de la anterior (un renglon por literal de
    Art OP). Se arrastra el expediente hacia abajo y se unen los conjuntos."""
    ws = openpyxl.load_workbook(path).worksheets[0]
    grupos, huerfanas, actual = {}, [], None
    for i in range(3, ws.max_row + 1):
        raw = ws.cell(i, REF["exp"]).value
        vals = {k: ws.cell(i, c).value for k, c in REF.items()}
        if all(v in (None, "") for v in vals.values()):
            continue
        if raw is None or str(raw).strip() == "":
            if actual is None:
                huerfanas.append(i)
                continue
            exp = actual                       # continuacion del anterior
        else:
            m = HYPERLINK.search(str(raw))
            exp = actual = norm_exp(m.group(1) if m else raw)
        g = grupos.setdefault(exp, dict(filas=[], vals={}, motivos=set(),
                                        art1=set(), art2=set(),
                                        fund1=set(), fund2=set()))
        g["filas"].append(i)
        for k, v in vals.items():             # primer valor no vacio gana
            if k not in g["vals"] or g["vals"][k] in (None, ""):
                g["vals"][k] = v
        g["motivos"] |= {norm_motivo(vals["mot1"]), norm_motivo(vals["mot2"])}
        g["art1"] |= set(norm_motivo(vals["art1"]).split(","))
        g["art2"] |= set(norm_motivo(vals["art2"]).split(","))
        g["fund1"].add(norm_si_no(vals["fund1"]))
        g["fund2"].add(norm_si_no(vals["fund2"]))
    for g in grupos.values():
        for k in ("motivos", "art1", "art2", "fund1", "fund2"):
            g[k] -= {""}
    return grupos, huerfanas


def leer_out(path):
    ws = openpyxl.load_workbook(path).worksheets[0]
    grupos = defaultdict(lambda: dict(filas=[], vals={}, motivos=set(), obs=set(),
                                      art1=set(), art2=set(),
                                      fund1=set(), fund2=set()))
    for i in range(3, ws.max_row + 1):
        raw = ws.cell(i, OUT["exp"]).value
        if raw is None or str(raw).strip() == "":
            continue
        exp = norm_exp(raw)
        vals = {k: ws.cell(i, c).value for k, c in OUT.items()}
        g = grupos[exp]
        g["filas"].append(i)
        for k, v in vals.items():
            if k not in g["vals"] or g["vals"][k] in (None, ""):
                g["vals"][k] = v
        g["motivos"].add(norm_motivo(vals["motivo"]))
        g["art1"] |= set(norm_motivo(vals["art1"]).split(","))
        g["art2"] |= set(norm_motivo(vals["art2"]).split(","))
        g["fund1"].add(norm_si_no(vals["fund1"]))
        g["fund2"].add(norm_si_no(vals["fund2"]))
        if norm(vals["obs"]):
            g["obs"].add(str(vals["obs"]).strip())
    for g in grupos.values():
        for k in ("motivos", "art1", "art2", "fund1", "fund2"):
            g[k] -= {""}
    return dict(grupos)


# ---------------------------------------------------------------- comparacion
# escalares: (clave, normalizador, etiqueta)
ESCALARES = [
    ("naturaleza", norm,       "Naturaleza"),
    ("oposicion",  norm_si_no, "Presenta Oposicion"),
    ("op1",        norm,       "Opositor 1"),
    ("op2",        norm,       "Opositor 2"),
    ("apelacion",  norm_si_no, "Apelacion"),
]
# conjuntos: (clave, etiqueta)
CONJUNTOS = [
    ("art1",    "Art OP 1"),
    ("fund1",   "Fundada OP 1"),
    ("art2",    "Art Opositor 2"),
    ("fund2",   "Fundada OP 2"),
    ("motivos", "Conjunto MOTIVOS"),
]


def _nuevo():
    return dict(ok=0, dif=0, vacio_ref=0, vacio_out=0, n=0,
                detalle=dict(igual=0, superconjunto=0, subconjunto=0,
                             solapado=0, disjunto=0))


def comparar(ref, out, apel_vacio_es_no=True):
    comunes = sorted(set(ref) & set(out))
    solo_ref = sorted(set(ref) - set(out))
    solo_out = sorted(set(out) - set(ref))

    metricas, disc = {}, []
    for k, fn, etiqueta in ESCALARES:
        m = metricas[etiqueta] = _nuevo()
        m["n"] = len(comunes)
        for e in comunes:
            a, b = fn(ref[e]["vals"][k]), fn(out[e]["vals"][k])
            if k == "apelacion" and apel_vacio_es_no and not b:
                b = "NO"     # nuestro programa nunca escribe NO: vacio == NO
            if a == b:
                m["ok"] += 1
            elif not a:
                m["vacio_ref"] += 1
                disc.append((e, etiqueta, a, b, "vacio_ref"))
            elif not b:
                m["vacio_out"] += 1
                disc.append((e, etiqueta, a, b, "vacio_out"))
            else:
                m["dif"] += 1
                disc.append((e, etiqueta, a, b, "distinto"))

    for k, etiqueta in CONJUNTOS:
        m = metricas[etiqueta] = _nuevo()
        m["n"] = len(comunes)
        d = m["detalle"]
        for e in comunes:
            a, b = ref[e][k], out[e][k]
            if not a and not b:
                m["ok"] += 1
                d["igual"] += 1
            elif not a:
                m["vacio_ref"] += 1
                disc.append((e, etiqueta, a, b, "vacio_ref"))
            elif not b:
                m["vacio_out"] += 1
                disc.append((e, etiqueta, a, b, "vacio_out"))
            elif a == b:
                m["ok"] += 1
                d["igual"] += 1
            else:
                tipo = ("superconjunto" if a < b else "subconjunto" if b < a
                        else "solapado" if a & b else "disjunto")
                m["dif"] += 1
                d[tipo] += 1
                disc.append((e, etiqueta, a, b, tipo))
    return comunes, solo_ref, solo_out, metricas, disc


def pct(x, n):
    return f"{100.0 * x / n:5.1f}%" if n else "  n/a"


def autotest():
    assert norm("  Sí  ") == "SI"
    assert norm_si_no("no") == norm_si_no("NO ") == "NO"
    m = HYPERLINK.search('=HYPERLINK("http://x?1","SD2022/0000017")')
    assert norm_exp(m.group(1)) == norm_exp("SD2022-0000017") == "SD2022-0000017"
    assert norm_motivo("136 a") == norm_motivo("Art. 136 literal a") == "136A"
    assert norm_motivo("136a,147") == "136A,147"
    assert norm("LABORATORIOS  LA SANTÉ  S.A.") == "LABORATORIOS LA SANTE S.A"
    print("autotest OK")


def main():
    if "--test" in sys.argv:
        autotest()
        return 0
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    out_path = args[0] if len(args) > 0 else OUT_DEFAULT
    ref_path = args[1] if len(args) > 1 else REF_DEFAULT

    ref, huerfanas = leer_ref(ref_path)
    out = leer_out(out_path)
    comunes, solo_ref, solo_out, met, disc = comparar(ref, out)

    print(f"REFERENCIA : {ref_path}")
    print(f"  expedientes : {len(ref)}   filas : {sum(len(g['filas']) for g in ref.values())}"
          f"   (de ellas {sum(len(g['filas']) - 1 for g in ref.values())} de continuacion)")
    if huerfanas:
        print(f"  filas huerfanas sin expediente previo: {huerfanas}")
    print(f"SALIDA     : {out_path}")
    print(f"  expedientes : {len(out)}   filas : {sum(len(g['filas']) for g in out.values())}")
    print()
    print("== 1. COBERTURA ==")
    print(f"  en ambos       : {len(comunes)}")
    print(f"  solo referencia: {len(solo_ref)}")
    for e in solo_ref:
        print(f"      {e}")
    print(f"  solo nuestra   : {len(solo_out)}")
    for e in solo_out:
        print(f"      {e}")
    print()
    print("== 2. COINCIDENCIA POR COLUMNA (sobre expedientes en ambos) ==")
    print("   (Apelacion: vacio nuestro se cuenta como NO; el programa nunca escribe NO)")
    print(f"{'Columna':<22}{'n':>6}{'coincide':>10}{'%':>8}"
          f"{'distinto':>10}{'%':>8}{'vacio ref':>11}{'vacio ntro':>12}")
    for k, m in met.items():
        n = m["n"]
        print(f"{k:<22}{n:>6}{m['ok']:>10}{pct(m['ok'], n):>8}"
              f"{m['dif']:>10}{pct(m['dif'], n):>8}"
              f"{m['vacio_ref']:>11}{m['vacio_out']:>12}")
    print()
    print("== 3. DETALLE DE LOS CAMPOS-CONJUNTO ==")
    print(f"{'Campo':<22}{'igual':>8}{'super':>8}{'sub':>8}{'solapado':>10}{'disjunto':>10}")
    for _, etiqueta in CONJUNTOS:
        d = met[etiqueta]["detalle"]
        print(f"{etiqueta:<22}{d['igual']:>8}{d['superconjunto']:>8}"
              f"{d['subconjunto']:>8}{d['solapado']:>10}{d['disjunto']:>10}")
    print()
    print("== 4. CRUCE CON Observaciones ==")
    exp_disc = {e for e, *_ in disc}
    print(f"  expedientes SIN ninguna discrepancia : {len(comunes) - len(exp_disc)}"
          f" / {len(comunes)}  ({pct(len(comunes) - len(exp_disc), len(comunes))})")
    marcados = {e for e in exp_disc if out[e]["obs"]}
    silenciosos = exp_disc - marcados
    print(f"  expedientes con alguna discrepancia : {len(exp_disc)}")
    print(f"  ...marcados en Observaciones        : {len(marcados)}")
    print(f"  ...SILENCIOSOS (sin observacion)    : {len(silenciosos)}")
    print(f"  expedientes con observacion y SIN discrepancia: "
          f"{len([e for e in comunes if out[e]['obs'] and e not in exp_disc])}")
    print()
    print("== DETALLE DE DISCREPANCIAS ==")
    for e, campo, a, b, tipo in sorted(disc, key=lambda x: (x[1], x[0])):
        mark = "OBS" if out[e]["obs"] else "   "
        print(f"  {mark} {e}  {campo:<20} {tipo:<14} ref={a!r:<45} ntro={b!r}")

    if "--csv" in sys.argv:
        import csv
        with open("discrepancias.csv", "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["expediente", "campo", "tipo", "ref", "nuestro",
                        "fila_ref", "filas_ntro", "observaciones"])
            for e, campo, a, b, tipo in sorted(disc, key=lambda x: (x[1], x[0])):
                w.writerow([e, campo, tipo, sorted(a) if isinstance(a, set) else a,
                            sorted(b) if isinstance(b, set) else b,
                            ref[e]["filas"], out[e]["filas"],
                            " | ".join(sorted(out[e]["obs"]))])
        print("\n-> discrepancias.csv escrito")
    return 0


if __name__ == "__main__":
    sys.exit(main())
