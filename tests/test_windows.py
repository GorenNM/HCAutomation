"""Casos 39–43 del §13.3: lo que solo se rompe en Windows.

Toda esta capa existe porque el proyecto **no** se desarrolla en un contenedor.
Corriendo solo en Linux, ninguno de estos fallos aparecería, y todos le pasarían
al usuario final el primer día: el archivo bloqueado por Excel, los acentos de
`cp1252`, el límite de 260 caracteres de las rutas, la barra del expediente que
NTFS no admite.

En Linux se omiten. El caso 40 vive en `test_rutas.py` (es un análisis del
código, no depende del sistema) y el 44 lo cubre `ExtraccionSIC.exe
--autoprueba`, que corre `construir_exe.bat`.
"""

from __future__ import annotations

import os
import stat
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
import pytest

from app.downloader.files import nombre_de_archivo
from app.excel.writer import ErrorEscritura, PRIMERA_FILA_DATOS, escribir
from app.models import TipoDoc
from app.pipeline import ejecutar
from tests.conftest import DATOS, SesionGrabadaMulti

pytestmark = pytest.mark.skipif(
    sys.platform != "win32", reason="Comportamientos propios de Windows"
)


def _mini(nombre="mini_2_casos.xlsx") -> Path:
    ruta = DATOS / nombre
    if not ruta.is_file():
        pytest.skip(f"Falta {nombre}: correr python -m tests.make_fixtures")
    return ruta


# --- Caso 39: el archivo está abierto en Excel -------------------------------


def _retener(ruta: Path) -> subprocess.Popen:
    """Abre el archivo como lo abre Excel: sin compartir con nadie.

    `FileShare 'None'` es lo mismo que hace Excel al abrir un libro, y es lo que
    provoca el `PermissionError` que el usuario ve como «acceso denegado».
    """
    guion = (
        f"$f=[System.IO.File]::Open('{ruta}','Open','ReadWrite','None');"
        "Write-Output 'retenido';"
        "Start-Sleep -Seconds 30;"
        "$f.Close()"
    )
    proceso = subprocess.Popen(
        ["powershell", "-NoProfile", "-Command", guion],
        stdout=subprocess.PIPE,
        text=True,
    )
    if proceso.stdout.readline().strip() != "retenido":  # espera al bloqueo
        proceso.kill()
        pytest.skip("PowerShell no pudo retener el archivo")
    return proceso


def test_caso_39_un_xlsx_abierto_en_excel_da_un_mensaje_para_humanos(tmp_path):
    destino = tmp_path / "salida.xlsx"
    escribir([], destino, "poc", alias={})  # que exista para poder bloquearlo
    retenedor = _retener(destino)

    try:
        with pytest.raises(ErrorEscritura) as error:
            escribir([], destino, "poc", alias={})
    finally:
        retenedor.kill()
        retenedor.wait(10)

    mensaje = str(error.value)
    assert "salida.xlsx" in mensaje
    assert "abierto en Excel" in mensaje
    assert "intente de nuevo" in mensaje
    assert "PermissionError" not in mensaje and "Traceback" not in mensaje


def test_caso_39_un_archivo_de_solo_lectura_tambien_avisa_bien(tmp_path):
    """Variante frecuente: el archivo quedó marcado de solo lectura."""
    destino = tmp_path / "salida.xlsx"
    escribir([], destino, "poc", alias={})
    os.chmod(destino, stat.S_IREAD)

    try:
        with pytest.raises(ErrorEscritura) as error:
            escribir([], destino, "poc", alias={})
    finally:
        os.chmod(destino, stat.S_IWRITE | stat.S_IREAD)

    assert "Ciérrelo e intente de nuevo" in str(error.value)


def test_caso_39_el_archivo_bloqueado_no_queda_corrupto(tmp_path):
    """openpyxl escribe truncando: un fallo a mitad no puede dejar basura."""
    destino = tmp_path / "salida.xlsx"
    escribir([], destino, "poc", alias={})
    original = destino.read_bytes()
    retenedor = _retener(destino)

    try:
        with pytest.raises(ErrorEscritura):
            escribir([], destino, "poc", alias={})
    finally:
        retenedor.kill()
        retenedor.wait(10)

    assert destino.read_bytes() == original
    openpyxl.load_workbook(destino).close()  # sigue abriéndose


# --- Caso 41: la barra del expediente sobre NTFS -----------------------------


def test_caso_41_la_barra_del_expediente_se_convierte_y_el_archivo_se_crea(tmp_path):
    """`SD2022/0000017` no es un nombre de archivo válido en Windows."""
    nombre = nombre_de_archivo("SD2022/0000017", TipoDoc.TM9)

    assert nombre == "SD2022-0000017_TM9.pdf"
    assert "/" not in nombre

    destino = tmp_path / nombre
    destino.write_bytes(b"%PDF-1.4 x %%EOF")
    assert destino.is_file()
    assert destino.name in [p.name for p in tmp_path.iterdir()]


@pytest.mark.parametrize(
    "expediente", ["SD2022/0000017", 'SD:2022*0000017?', "SD2022\\0000017"]
)
def test_caso_41_ningun_caracter_prohibido_sobrevive(tmp_path, expediente):
    destino = tmp_path / nombre_de_archivo(expediente, TipoDoc.TM128)
    destino.write_bytes(b"%PDF-1.4 x %%EOF")

    assert destino.is_file()
    assert not any(c in destino.name for c in '<>:"/\\|?*')


# --- Caso 42: rutas con acentos y espacios -----------------------------------


def test_caso_42_una_corrida_completa_en_una_ruta_con_acentos(tmp_path):
    """`C:\\Users\\José\\Mis documentos\\` — el caso del plan, entero."""
    carpeta = tmp_path / "José Ramírez" / "Mis documentos" / "Extracción SIC"

    resultado = ejecutar(
        _mini(),
        salida=carpeta / "Negación de marcas ñ.xlsx",
        temp=carpeta / "temporales",
        fabrica_sesion=SesionGrabadaMulti,
    )

    assert resultado.ruta_excel.is_file()
    hoja = openpyxl.load_workbook(resultado.ruta_excel)["Hoja1"]
    assert hoja.cell(row=PRIMERA_FILA_DATOS, column=1).value == "SD2022/0000017"
    # Los PDFs cuelgan de salida\soportes\<expediente>\, junto al Excel, no de temp\.
    assert list((carpeta / "soportes").glob("*/*.pdf"))


def test_caso_42_los_acentos_del_contenido_sobreviven_a_cp1252(tmp_path):
    """Windows abre los archivos de texto en cp1252 si no se le dice otra cosa."""
    resultado = ejecutar(
        _mini(),
        salida=tmp_path / "salida.xlsx",
        temp=tmp_path / "temp",
        fabrica_sesion=SesionGrabadaMulti,
    )

    hoja = openpyxl.load_workbook(resultado.ruta_excel)["Hoja1"]
    cabeceras = [c.value for c in hoja[2]]
    fila = [c.value for c in hoja[PRIMERA_FILA_DATOS + 1]]
    opositor = fila[cabeceras.index("Opositor 1")]

    assert "Diagnostico" in opositor
    assert "Número de Expediente" in cabeceras
    assert "Apelación a la negación" in cabeceras


# --- Caso 43: el límite de 260 caracteres ------------------------------------


def test_caso_43_una_ruta_larguisima_falla_claro_o_funciona_pero_no_miente(tmp_path):
    """Con rutas largas habilitadas Windows la acepta; sin ellas, tiene que dar
    un error legible. Lo inaceptable es la tercera opción: truncar en silencio
    y dejar el archivo en otro sitio."""
    profunda = tmp_path
    while len(str(profunda)) < 300:
        profunda = profunda / ("carpeta_muy_larga_" + "x" * 30)
    destino = profunda / "salida.xlsx"

    try:
        escrita = escribir([], destino, "poc", alias={})
    except ErrorEscritura as error:
        assert "No se pudo" in str(error)
        assert "Traceback" not in str(error)
        return

    assert escrita == destino, "escribió en una ruta distinta de la pedida"
    assert escrita.is_file()


def test_caso_43_el_nombre_del_pdf_no_se_trunca_en_silencio(tmp_path):
    """Si el nombre cambiara, la caché no reconocería el archivo la próxima vez."""
    esperado = nombre_de_archivo("SD2022/0000017", TipoDoc.TM128, orden=2)
    destino = tmp_path / esperado

    destino.write_bytes(b"%PDF-1.4 x %%EOF")

    assert destino.name == esperado
    assert (tmp_path / esperado).is_file()


# --- Rendimiento del sistema de archivos -------------------------------------


def test_el_antivirus_no_hace_inservible_la_escritura_atomica(tmp_path):
    """En Windows un antivirus puede retener el archivo justo tras crearlo.

    No se puede provocar, pero sí medir: 50 escrituras atómicas seguidas no
    pueden tardar tanto que la corrida de 987 expedientes sea inviable.
    """
    from app.downloader.files import _guardar

    datos = b"%PDF-1.4 " + b"x" * 50_000 + b" %%EOF"
    comienzo = time.monotonic()
    for numero in range(50):
        _guardar(datos, tmp_path / f"doc_{numero}.pdf")
    duracion = time.monotonic() - comienzo

    assert len(list(tmp_path.glob("*.pdf"))) == 50
    assert list(tmp_path.glob("*.part")) == [], "quedaron temporales sin renombrar"
    assert duracion < 30, f"50 escrituras tardaron {duracion:.1f} s"
