"""Lectura del Excel de entrada (el reporte exportado desde SIPI).

El detalle que define este módulo: la primera columna **no tiene hyperlink**.
Tiene una fórmula de Excel:

    =HYPERLINK("http://sipi.sic.gov.co/sipi/View.ashx?3857028","SD2022/0000017")

`cell.hyperlink` devuelve `None`, así que el workbook se abre con
`data_only=False` y la URL y el expediente se sacan de la fórmula con un regex.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from app.config import FILA_CABECERAS
from app.models import SourceRow
from app.utils.text import normalizar, sin_tildes

log = logging.getLogger(__name__)

# =HYPERLINK("<url>","<expediente>")  — las comillas internas nunca aparecen
# escapadas en los reportes reales de SIPI.
_FORMULA_HYPERLINK = re.compile(
    r'^\s*=\s*HYPERLINK\s*\(\s*"([^"]+)"\s*,\s*"([^"]+)"\s*\)\s*$', re.IGNORECASE
)

# SD2022/0000017 — dos o tres letras, año, barra, consecutivo.
_EXPEDIENTE = re.compile(r"\b([A-Z]{2,3}\d{4}/\d{4,9})\b")

_URL_VALIDA = re.compile(r"sipi\.sic\.gov\.co", re.IGNORECASE)

# Cabecera del reporte -> campo de SourceRow.
_COLUMNAS = {
    "numero de caso": "enlace",
    "caso titulo": "marca",
    "titular": "titular",
    "descripcion de productos y servicios": "niza",
    "productos y servicios descripcion": "descripcion",
    "bajo oposicion": "bajo_oposicion",
}


class ErrorLectura(Exception):
    """El archivo no se puede leer. El mensaje va directo al usuario."""


@dataclass
class ResultadoLectura:
    filas: list[SourceRow] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    @property
    def expedientes_unicos(self) -> set[str]:
        return {f.expediente for f in self.filas}


def _clave(cabecera: object) -> str:
    """Normaliza una cabecera para compararla sin depender de tildes ni espacios."""
    return sin_tildes(normalizar(str(cabecera or ""))).casefold()


def _mapear_columnas(cabeceras: tuple[object, ...]) -> dict[str, int]:
    """Ubica cada columna por su nombre, no por su posición.

    Si mañana SIPI inserta una columna, esto sigue funcionando; si le cambia el
    nombre a una que necesitamos, falla con un mensaje que dice cuál.
    """
    encontradas: dict[str, int] = {}
    for indice, cabecera in enumerate(cabeceras):
        campo = _COLUMNAS.get(_clave(cabecera))
        if campo is not None and campo not in encontradas:
            encontradas[campo] = indice

    faltantes = sorted(set(_COLUMNAS.values()) - set(encontradas))
    if faltantes:
        vistas = [str(c) for c in cabeceras if c]
        raise ErrorLectura(
            "El Excel no tiene el formato del reporte de SIPI: faltan las columnas "
            f"{faltantes}. Se esperaban las cabeceras en la fila {FILA_CABECERAS}. "
            f"Se encontró: {vistas[:6]}"
        )
    return encontradas


def _abrir(ruta: Path):
    try:
        return openpyxl.load_workbook(ruta, read_only=True, data_only=False)
    except FileNotFoundError as exc:
        raise ErrorLectura(f"No se encontró el archivo «{ruta}».") from exc
    except (InvalidFileException, BadZipFile) as exc:
        # Un .xlsx es un zip. Si no lo es, el archivo está renombrado o es de
        # otro formato: BadZipFile llega antes de que openpyxl opine.
        raise ErrorLectura(
            f"«{ruta.name}» no es un archivo de Excel (.xlsx) válido. "
            "Si viene de un .xls antiguo, ábralo y guárdelo como .xlsx."
        ) from exc
    except (OSError, KeyError, ValueError, TypeError) as exc:
        # openpyxl lanza cosas variadas ante un zip corrupto; el usuario no
        # tiene por qué ver una traza de xml.
        raise ErrorLectura(
            f"No se pudo abrir «{ruta.name}»: el archivo parece dañado ({exc})."
        ) from exc


def _texto(valor: object) -> str:
    return normalizar(str(valor)) if valor is not None else ""


def _parsear_enlace(celda: object) -> tuple[str, str] | None:
    """Devuelve (expediente, url) o None si la celda no sirve.

    La url puede venir vacía: hay celdas con el expediente en texto plano, sin
    fórmula. Esa fila conserva sus datos del Excel aunque no se pueda descargar.
    """
    crudo = _texto(celda)
    if not crudo:
        return None

    formula = _FORMULA_HYPERLINK.match(crudo)
    if formula:
        url, etiqueta = formula.group(1), formula.group(2)
        expediente = _EXPEDIENTE.search(etiqueta.upper())
        if expediente and _URL_VALIDA.search(url):
            return expediente.group(1), url

    # Fórmula rota, o texto plano: se rescata el expediente si está ahí.
    suelto = _EXPEDIENTE.search(crudo.upper())
    if suelto:
        return suelto.group(1), ""
    return None


def _filas_de_datos(hoja) -> Iterator[tuple[int, tuple[object, ...]]]:
    for numero, valores in enumerate(
        hoja.iter_rows(min_row=FILA_CABECERAS + 1, values_only=True),
        start=FILA_CABECERAS + 1,
    ):
        yield numero, valores


def leer_reporte(ruta: str | Path) -> ResultadoLectura:
    """Lee el reporte y devuelve una fila por expediente, más los avisos."""
    ruta = Path(ruta)
    libro = _abrir(ruta)
    try:
        hoja = libro.worksheets[0]
        cabeceras = next(
            hoja.iter_rows(
                min_row=FILA_CABECERAS, max_row=FILA_CABECERAS, values_only=True
            ),
            None,
        )
        if cabeceras is None:
            raise ErrorLectura(
                f"«{ruta.name}» no tiene datos: se esperaban las cabeceras en la "
                f"fila {FILA_CABECERAS}."
            )
        columnas = _mapear_columnas(cabeceras)
        resultado = _recorrer(hoja, columnas)
    finally:
        libro.close()

    _avisar_duplicados(resultado)
    log.info(
        "Leídas %d filas de «%s» (%d expedientes únicos, %d avisos)",
        len(resultado.filas),
        ruta.name,
        len(resultado.expedientes_unicos),
        len(resultado.avisos),
    )
    return resultado


def _recorrer(hoja, columnas: dict[str, int]) -> ResultadoLectura:
    resultado = ResultadoLectura()
    ancho_minimo = max(columnas.values()) + 1

    for numero, valores in _filas_de_datos(hoja):
        if not any(v not in (None, "") for v in valores):
            continue  # fila en blanco: separadores y pie del reporte
        if len(valores) < ancho_minimo:
            valores = valores + (None,) * (ancho_minimo - len(valores))

        enlace = _parsear_enlace(valores[columnas["enlace"]])
        if enlace is None:
            resultado.avisos.append(
                f"Fila {numero}: se omite, no contiene un número de expediente "
                f"reconocible ({_texto(valores[columnas['enlace']])[:60]!r})."
            )
            continue

        expediente, url = enlace
        if not url:
            resultado.avisos.append(
                f"Fila {numero}: el expediente {expediente} no trae enlace; "
                "no se podrán descargar sus documentos."
            )

        resultado.filas.append(
            SourceRow(
                expediente=expediente,
                case_url=url,
                marca=_texto(valores[columnas["marca"]]),
                titular=_texto(valores[columnas["titular"]]),
                niza=_texto(valores[columnas["niza"]]),
                descripcion=_texto(valores[columnas["descripcion"]]),
                bajo_oposicion=_texto(valores[columnas["bajo_oposicion"]]),
                fila=numero,
            )
        )
    return resultado


def _avisar_duplicados(resultado: ResultadoLectura) -> None:
    vistos: dict[str, int] = {}
    for fila in resultado.filas:
        if fila.expediente in vistos:
            resultado.avisos.append(
                f"Fila {fila.fila}: el expediente {fila.expediente} ya aparecía en la "
                f"fila {vistos[fila.expediente]}; se descargará una sola vez."
            )
        else:
            vistos[fila.expediente] = fila.fila
