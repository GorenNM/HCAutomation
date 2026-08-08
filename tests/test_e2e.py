"""De punta a punta con datos reales, sin red.

Los tres Excel de `tests/data/` son **recortes del reporte de verdad**: se
abrieron con openpyxl y se borraron filas, así que la columna A trae la misma
fórmula `HYPERLINK`, las mismas cabeceras en la fila 11 y las mismas rarezas de
formato que el archivo que usa el usuario. Las respuestas de SIPI son las
grabadas en `tests/fixtures/http/`.

Lo que se comprueba aquí no es que cada módulo funcione —de eso se encargan las
otras pruebas— sino que **encajan**: del Excel que llega al Excel que sale, sin
que nadie toque nada por el camino.
"""

from __future__ import annotations

import openpyxl
import pytest

from app.excel.writer import CABECERAS, PRIMERA_FILA_DATOS
from app.pipeline import ejecutar
from tests.conftest import DATOS, SesionGrabadaMulti

pytestmark = pytest.mark.usefixtures("hay_fixtures_http")


def corrida(nombre: str, tmp_path, **kwargs):
    entrada = DATOS / nombre
    if not entrada.is_file():
        pytest.skip(f"Falta {nombre}: correr python -m tests.make_fixtures")
    kwargs.setdefault("hilos", 2)
    kwargs.setdefault("fabrica_sesion", SesionGrabadaMulti)
    return ejecutar(
        entrada, salida=tmp_path / "salida.xlsx", temp=tmp_path / "temp", **kwargs
    )


def hoja_de(ruta):
    return openpyxl.load_workbook(ruta)["Hoja1"]


def filas_como_diccionarios(ruta) -> list[dict]:
    hoja = hoja_de(ruta)
    cabeceras = [celda.value for celda in hoja[2]]
    return [
        dict(zip(cabeceras, [celda.value for celda in fila]))
        for fila in hoja.iter_rows(min_row=PRIMERA_FILA_DATOS)
    ]


# --- Un caso -----------------------------------------------------------------


def test_un_expediente_sin_oposicion_de_principio_a_fin(tmp_path):
    resultado = corrida("mini_1_caso.xlsx", tmp_path)

    (fila,) = filas_como_diccionarios(resultado.ruta_excel)
    assert fila["Número de Expediente"] == "SD2022/0000017"
    assert fila["Marca"] == "MALTAVITAN"
    assert fila["Naturaleza"] == "Nominativa"
    assert fila["Presenta Oposición"] == "No"
    assert fila["MOTIVO Negación"] == "136a"
    assert fila["Apelación a la negación"] == "SI"
    assert fila["Opositor 1"] is None
    assert fila["Observaciones"] is None, "no debería haber nada que revisar a mano"
    assert fila["NIZA"] == "5"
    assert "farmacéuticos" in fila["Descripción de Productos y Servicios"]


# --- Dos casos ---------------------------------------------------------------


def test_dos_expedientes_con_y_sin_oposicion(tmp_path):
    resultado = corrida("mini_2_casos.xlsx", tmp_path)

    filas = filas_como_diccionarios(resultado.ruta_excel)
    assert [f["Número de Expediente"] for f in filas] == [
        "SD2022/0000017",
        "SD2022/0001545",
    ]

    con_oposicion = filas[1]
    assert con_oposicion["Presenta Oposición"] == "Sí"
    assert con_oposicion["Opositor 1"] == "Grupo Diagnostico S.A. Dimed S.A."
    assert con_oposicion["Art OP 1"] == "136a, 136b"
    assert con_oposicion["Fundada OP 1"] == "SI"
    assert con_oposicion["MOTIVO Negación"] == "136a"


def test_el_caso_de_la_negacion_lexica_llega_intacto_hasta_el_excel(tmp_path):
    """El riesgo número uno del proyecto, comprobado en la salida final.

    La resolución dice «NO está comprendido en … literal b)». Si 136b apareciera
    como motivo, el sistema estaría afirmando lo contrario que el documento.
    """
    resultado = corrida("mini_2_casos.xlsx", tmp_path)

    fila = filas_como_diccionarios(resultado.ruta_excel)[1]
    assert fila["MOTIVO Negación"] == "136a"
    assert "136b" not in (fila["MOTIVO Negación"] or "")
    assert "136b" in fila["Observaciones"], "el descarte tiene que quedar explicado"
    assert "NO está comprendido" in fila["Observaciones"]


# --- Varios motivos ----------------------------------------------------------


def test_un_expediente_con_dos_motivos_da_dos_filas(tmp_path):
    """El requisito que motivó todo el cambio de formato."""
    resultado = corrida("mini_multimotivo.xlsx", tmp_path)

    filas = filas_como_diccionarios(resultado.ruta_excel)
    assert len(filas) == 2
    assert [f["MOTIVO Negación"] for f in filas] == ["136a", "136h"]
    assert [f["Motivo #"] for f in filas] == ["1", "2"]
    assert [f["Motivos totales"] for f in filas] == ["2", "2"]


def test_todo_lo_demas_se_repite_en_las_dos_filas(tmp_path):
    """«Repitiendo el resto de la información», textual del requisito."""
    resultado = corrida("mini_multimotivo.xlsx", tmp_path)

    primera, segunda = filas_como_diccionarios(resultado.ruta_excel)
    distintas = {
        clave
        for clave in primera
        if primera[clave] != segunda[clave]
    }
    assert distintas == {"MOTIVO Negación", "Motivo #"}


def test_los_dos_opositores_salen_en_sus_columnas(tmp_path):
    resultado = corrida("mini_multimotivo.xlsx", tmp_path)

    fila = filas_como_diccionarios(resultado.ruta_excel)[0]
    assert fila["Opositor 1"] == "SOCIETE DES PRODUITS NESTLE SA"
    assert fila["Opositor 2"] == "KRAFT FOODS SCHWEIZ HOLDING GMBH"
    assert fila["Fundada OP 1"] == "SI"
    assert fila["Fundada OP 2"] == "SI"


# --- Estructura del archivo generado -----------------------------------------


def test_el_archivo_generado_tiene_la_forma_de_la_referencia(tmp_path):
    resultado = corrida("mini_2_casos.xlsx", tmp_path)
    hoja = hoja_de(resultado.ruta_excel)

    assert [celda.value for celda in hoja[2]] == list(CABECERAS["poc"])
    assert {str(r) for r in hoja.merged_cells.ranges} == {"D1:H1", "I1:L1"}
    assert hoja.freeze_panes == "A3"
    assert hoja.cell(row=PRIMERA_FILA_DATOS, column=1).hyperlink is not None


def test_el_excel_generado_se_puede_volver_a_abrir_con_excel(tmp_path):
    """Ida y vuelta completa: si openpyxl lo relee, Excel también lo abre."""
    resultado = corrida("mini_2_casos.xlsx", tmp_path)

    libro = openpyxl.load_workbook(resultado.ruta_excel, data_only=True)
    try:
        assert libro.sheetnames == ["Hoja1"]
        assert libro["Hoja1"].max_row == PRIMERA_FILA_DATOS + 1
    finally:
        libro.close()


# --- Reanudación -------------------------------------------------------------


def test_repetir_la_corrida_no_vuelve_a_descargar_ni_cambia_el_resultado(tmp_path):
    primera = corrida("mini_2_casos.xlsx", tmp_path)
    filas_primera = filas_como_diccionarios(primera.ruta_excel)

    segunda = ejecutar(
        DATOS / "mini_2_casos.xlsx",
        salida=tmp_path / "otra.xlsx",
        temp=tmp_path / "temp",
        hilos=2,
        fabrica_sesion=SesionGrabadaMulti,
    )

    assert primera.progreso.pdfs > 0
    assert segunda.progreso.pdfs == 0
    assert filas_como_diccionarios(segunda.ruta_excel) == filas_primera


def test_un_pdf_corrupto_en_la_cache_se_vuelve_a_bajar(tmp_path):
    """La caché no puede envenenar la siguiente corrida."""
    primera = corrida("mini_1_caso.xlsx", tmp_path)
    victima = tmp_path / "soportes" / "SD2022-0000017" / "SD2022-0000017_TM9.pdf"
    victima.write_bytes(b"%PDF- pero cortado por la mitad")

    segunda = ejecutar(
        DATOS / "mini_1_caso.xlsx",
        salida=tmp_path / "otra.xlsx",
        temp=tmp_path / "temp",
        fabrica_sesion=SesionGrabadaMulti,
    )

    assert segunda.progreso.pdfs == 1, "no se rebajó el PDF roto"
    assert filas_como_diccionarios(segunda.ruta_excel) == filas_como_diccionarios(
        primera.ruta_excel
    )


# --- Sin red -----------------------------------------------------------------


def test_sin_red_sale_un_excel_con_las_filas_y_el_motivo_del_fallo(tmp_path):
    class SinRed(SesionGrabadaMulti):
        def obtener(self, url, referer=None):
            raise OSError("no hay conexión")

    resultado = corrida("mini_2_casos.xlsx", tmp_path, fabrica_sesion=SinRed)

    filas = filas_como_diccionarios(resultado.ruta_excel)
    assert len(filas) == 2, "ningún expediente puede desaparecer de la salida"
    assert all(f["MOTIVO Negación"] is None for f in filas)
    assert all("conexión" in f["Observaciones"] for f in filas)
    assert resultado.progreso.errores == 2
