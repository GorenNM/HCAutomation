"""Genera las fixtures de prueba a partir de los archivos reales.

Se ejecuta a mano, una sola vez (o cuando cambien los datos de origen):

    python -m tests.make_fixtures

Los mini-Excel son **recortes del reporte real**, no archivos sintéticos:
conservan las 11 filas de cabecera, los estilos y las fórmulas HYPERLINK tal
como las exporta SIPI. Si se generaran a mano, el E2E dejaría de probar el
lector real y pasaría a probar nuestra idea de cómo es el archivo.
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[1]
ORIGEN = RAIZ / "Reporte 5 Enero 2025.xlsx"
DESTINO = RAIZ / "tests" / "data"

PRIMERA_FILA_DATOS = 12

# Expedientes elegidos por lo que ejercitan, no por comodidad:
RECORTES: dict[str, list[str]] = {
    # Sin oposición, un solo motivo (136a).
    "mini_1_caso.xlsx": ["SD2022/0000017"],
    # Añade el caso de la negación léxica: causal a) aplicada, b) descartada.
    "mini_2_casos.xlsx": ["SD2022/0000017", "SD2022/0001545"],
    # Expediente con dos motivos de negación -> debe producir dos filas.
    "mini_multimotivo.xlsx": ["SD2022/0097089"],
}

_EXPEDIENTE_EN_FORMULA = re.compile(r'"([A-Z]{2,3}\d{4}/\d{4,9})"\s*\)')


def _expediente_de(celda: object) -> str | None:
    encontrado = _EXPEDIENTE_EN_FORMULA.search(str(celda or ""))
    return encontrado.group(1) if encontrado else None


def recortar(origen: Path, destino: Path, expedientes: list[str]) -> int:
    """Copia el reporte dejando solo las filas de `expedientes`."""
    shutil.copyfile(origen, destino)
    libro = openpyxl.load_workbook(destino)
    hoja = libro.worksheets[0]

    conservar = set(expedientes)
    a_borrar = [
        fila
        for fila in range(PRIMERA_FILA_DATOS, hoja.max_row + 1)
        if _expediente_de(hoja.cell(row=fila, column=1).value) not in conservar
    ]
    for fila in reversed(a_borrar):  # de abajo hacia arriba: los índices no se corren
        hoja.delete_rows(fila)

    libro.save(destino)
    libro.close()
    return hoja.max_row - PRIMERA_FILA_DATOS + 1


# --- Grabación de las respuestas reales de SIPI ------------------------------
#
# A partir de aquí la suite es offline: se guardan las respuestas tal cual las
# devuelve el sitio y se reproducen con `SesionGrabada` (ver tests/conftest.py).
# Incluido el PNG de 828 bytes que SIPI manda cuando falta la cookie de sesión:
# esa respuesta es una fixture de primera clase, no un descarte.

HTTP = RAIZ / "tests" / "fixtures" / "http"

EXPEDIENTES_A_GRABAR = [
    "SD2022/0000017",  # sin oposición, 3 documentos
    "SD2022/0001545",  # con oposición y negación léxica, 5 documentos
    "SD2022/0097089",  # dos motivos de negación
]


def _url_del_expediente(expediente: str) -> str:
    """Busca en el reporte real la URL de un expediente."""
    libro = openpyxl.load_workbook(ORIGEN, read_only=True, data_only=False)
    try:
        for fila in libro.worksheets[0].iter_rows(min_row=PRIMERA_FILA_DATOS,
                                                  values_only=True):
            if _expediente_de(fila[0]) == expediente:
                url = re.search(r'"([^"]+)"', str(fila[0]))
                if url:
                    return url.group(1)
    finally:
        libro.close()
    raise SystemExit(f"{expediente} no está en el reporte")


def _carpeta(expediente: str) -> Path:
    return HTTP / expediente.replace("/", "-")


def grabar_http(forzar: bool = False) -> None:
    """Descarga y guarda en disco el HTML y los PDFs de unos pocos expedientes."""
    import json

    from app.downloader.scraper import documentos_de_expediente
    from app.downloader.session import SesionSIPI

    HTTP.mkdir(parents=True, exist_ok=True)

    with SesionSIPI() as sesion:
        for expediente in EXPEDIENTES_A_GRABAR:
            carpeta = _carpeta(expediente)
            if carpeta.is_dir() and not forzar:
                print(f"{expediente}: ya grabado, se omite (--forzar para rehacer)")
                continue
            carpeta.mkdir(parents=True, exist_ok=True)

            case_url = _url_del_expediente(expediente)
            pagina, documentos = documentos_de_expediente(sesion, case_url)
            (carpeta / "browse.html").write_text(pagina.html, encoding="utf-8")

            indice = {"case_url": case_url, "url_final": pagina.url, "documentos": []}
            for numero, documento in enumerate(documentos, 1):
                respuesta = sesion.obtener(documento.url, referer=pagina.url)
                nombre = f"doc{numero:02d}.bin"
                (carpeta / nombre).write_bytes(respuesta.content)
                indice["documentos"].append(
                    {
                        "archivo": nombre,
                        "url": documento.url,
                        "etiqueta": documento.etiqueta,
                        "tipo": documento.tipo.value,
                        "content_type": respuesta.headers.get("Content-Type", ""),
                        "bytes": len(respuesta.content),
                    }
                )
            (carpeta / "indice.json").write_text(
                json.dumps(indice, indent=2, ensure_ascii=False), encoding="utf-8"
            )
            print(f"{expediente}: {len(documentos)} documento(s) grabados")

    _grabar_respuesta_sin_sesion()


def _grabar_respuesta_sin_sesion() -> None:
    """El PNG de 828 bytes: qué devuelve GetFile.aspx sin cookie de sesión.

    Ojo: **no todos** los documentos la exigen. Los anexos y las apelaciones
    suelen bajar sin cookie; el TM9/TM128 no. Por eso se prueban en orden hasta
    dar con uno que falle, en vez de asumir el primero.
    """
    import json

    import requests

    from app.downloader.session import forzar_https

    for expediente in EXPEDIENTES_A_GRABAR:
        referencia = _carpeta(expediente) / "indice.json"
        if not referencia.is_file():
            continue
        for documento in json.loads(referencia.read_text(encoding="utf-8"))["documentos"]:
            respuesta = requests.get(forzar_https(documento["url"]), timeout=60)
            if respuesta.content.startswith(b"%PDF-"):
                continue  # este no necesita sesión: no sirve como fixture
            destino = HTTP / "sin_sesion.bin"
            destino.write_bytes(respuesta.content)
            print(
                f"sin_sesion.bin: {len(respuesta.content)} bytes, "
                f"{respuesta.headers.get('Content-Type')} "
                f"(HTTP {respuesta.status_code}) — de {documento['tipo']}"
            )
            return
    print("AVISO: ningún documento falló sin sesión; no se grabó sin_sesion.bin")


def grabar_texto() -> None:
    """Guarda el texto ya extraído de las resoluciones.

    Así las pruebas del extractor no dependen de pypdf ni de leer 11 PDFs en
    cada corrida: lo que se prueba ahí es el análisis del texto, no la lectura.
    """
    import json

    from app.parser.pdf_text import texto_de_pdf

    destino = RAIZ / "tests" / "fixtures" / "texto"
    destino.mkdir(parents=True, exist_ok=True)

    for expediente in EXPEDIENTES_A_GRABAR:
        carpeta = _carpeta(expediente)
        indice = carpeta / "indice.json"
        if not indice.is_file():
            continue
        for documento in json.loads(indice.read_text(encoding="utf-8"))["documentos"]:
            if documento["tipo"] not in ("TM9", "TM128"):
                continue
            texto = texto_de_pdf(carpeta / documento["archivo"])
            archivo = destino / f"{carpeta.name}_{documento['tipo']}.txt"
            archivo.write_text(texto, encoding="utf-8")
            print(f"{archivo.name}: {len(texto)} caracteres")


def main() -> None:
    import sys

    if not ORIGEN.is_file():
        raise SystemExit(f"No está el reporte de origen: {ORIGEN}")
    forzar = "--forzar" in sys.argv
    solo_http = "--http" in sys.argv

    if not solo_http:
        DESTINO.mkdir(parents=True, exist_ok=True)
        for nombre, expedientes in RECORTES.items():
            cuantas = recortar(ORIGEN, DESTINO / nombre, expedientes)
            print(f"{nombre}: {cuantas} fila(s) — {', '.join(expedientes)}")

    grabar_http(forzar=forzar)
    grabar_texto()


if __name__ == "__main__":
    main()
