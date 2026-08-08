"""Pruebas del sembrador de `alias.json` y del diccionario que produce."""

from __future__ import annotations

import json

import openpyxl
import pytest

from app.excel.writer import cargar_alias, expandir, valores
from app.models import ExtractedData, Opositor, SourceRow
from app.utils.text import clave_comparacion
from sembrar_alias import PARES_DE_COLUMNAS, PRIMERA_FILA, construir, leer_pares
from tests.conftest import RAIZ

REFERENCIA = RAIZ / "Negacion marcas con información extra.xlsx"
ALIAS = RAIZ / "alias.json"


def fuente() -> SourceRow:
    return SourceRow("SD2022/0000017", "", "MARCA", "TIT", "5", "desc", "No", 12)


def corto_de(nombre: str, alias: dict[str, str]) -> str:
    datos = ExtractedData(motivos=["136a"], opositores=[Opositor(nombre)])
    return valores(expandir(fuente(), datos)[0], alias, "poc")[5]


def construir_referencia(tmp_path, filas):
    """Un archivo con la forma del de referencia: cabeceras en la fila 2."""
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append([None] * 18)
    hoja.append(["Número de Expediente"] + [None] * 17)
    for nombre1, corto1, nombre2, corto2 in filas:
        fila = [None] * 18
        fila[4], fila[5], fila[8], fila[9] = nombre1, corto1, nombre2, corto2
        hoja.append(fila)
    ruta = tmp_path / "referencia.xlsx"
    libro.save(ruta)
    libro.close()
    return ruta


# --- Lectura del archivo de referencia ----------------------------------------


def test_las_columnas_son_las_del_archivo_real():
    if not REFERENCIA.is_file():
        pytest.skip("No está el archivo de referencia")
    hoja = openpyxl.load_workbook(REFERENCIA, read_only=True)["Hoja1"]
    cabeceras = [c.value for c in next(hoja.iter_rows(min_row=PRIMERA_FILA - 1))]

    for col_nombre, col_corto in PARES_DE_COLUMNAS:
        assert "positor" in str(cabeceras[col_nombre])
        assert "Nombre corto" in str(cabeceras[col_corto])


def test_del_archivo_real_salen_mas_de_cien_opositores():
    if not REFERENCIA.is_file():
        pytest.skip("No está el archivo de referencia")
    alias, _avisos = construir(leer_pares(REFERENCIA))
    assert len(alias) > 100


def test_una_fila_sin_nombre_corto_no_entra(tmp_path):
    ruta = construir_referencia(
        tmp_path, [("ALFA S.A.", None, None, None), ("BETA LTDA", "Beta", None, None)]
    )
    alias, _ = construir(leer_pares(ruta))

    assert clave_comparacion("BETA LTDA") in alias
    assert clave_comparacion("ALFA S.A.") not in alias


def test_los_saltos_de_linea_dentro_del_nombre_no_parten_la_clave(tmp_path):
    """En el archivo real hay nombres con un `\\n` dentro."""
    ruta = construir_referencia(
        tmp_path, [("LABORATORIOS\nLEGRAND S.A.", "Legrand", None, None)]
    )
    alias, _ = construir(leer_pares(ruta))

    assert corto_de("Laboratorios Legrand S.A.", alias) == "Legrand"


def test_el_segundo_opositor_tambien_cuenta(tmp_path):
    ruta = construir_referencia(tmp_path, [(None, None, "GAMMA INC", "Gamma")])
    alias, _ = construir(leer_pares(ruta))

    assert corto_de("GAMMA INC", alias) == "Gamma"


# --- Conflictos ---------------------------------------------------------------


def test_dos_nombres_cortos_para_el_mismo_opositor_avisan(tmp_path):
    """Erratas reales del archivo: `Nureo` / `Nutreo`, `Legrand` / `LabLegrand`."""
    ruta = construir_referencia(
        tmp_path,
        [
            ("C.I. NUTREO S.A.S.", "Nutreo", None, None),
            ("C.I. NUTREO S.A.S.", "Nutreo", None, None),
            ("C.I. NUTREO S.A.S.", "Nureo", None, None),
        ],
    )
    alias, avisos = construir(leer_pares(ruta))

    assert alias[clave_comparacion("C.I. NUTREO S.A.S.")] == "Nutreo", "gana el frecuente"
    assert len(avisos) == 1
    assert "Nureo" in avisos[0] and "Nutreo" in avisos[0]


def test_el_empate_se_resuelve_igual_en_las_dos_corridas(tmp_path):
    ruta = construir_referencia(
        tmp_path,
        [("X S.A.", "Uno", None, None), ("X S.A.", "Dos", None, None)],
    )
    primera, _ = construir(leer_pares(ruta))
    segunda, _ = construir(leer_pares(ruta))

    assert primera == segunda


def test_sin_conflictos_no_hay_avisos(tmp_path):
    ruta = construir_referencia(
        tmp_path, [("ALFA S.A.", "Alfa", "BETA LTDA", "Beta")]
    )
    _alias, avisos = construir(leer_pares(ruta))

    assert avisos == []


def test_dos_empresas_distintas_no_se_funden(tmp_path):
    """`MUÑOZ` y `MUNOZ` son distintas: la ñ no es una n acentuada."""
    ruta = construir_referencia(
        tmp_path, [("MUÑOZ S.A.", "Munioz", "MUNOZ S.A.", "Munoz")]
    )
    alias, avisos = construir(leer_pares(ruta))

    assert len(alias) == 2
    assert avisos == []


def test_las_variantes_de_escritura_si_se_funden(tmp_path):
    """El mismo opositor escrito con y sin tildes es uno solo."""
    ruta = construir_referencia(
        tmp_path,
        [
            ("Grupo Diagnóstico S.A. Dimed S.A.", "Dimed", None, None),
            ("GRUPO DIAGNOSTICO DIMED", "Dimed", None, None),
        ],
    )
    alias, avisos = construir(leer_pares(ruta))

    assert len(alias) == 1
    assert avisos == []


# --- El archivo generado ------------------------------------------------------


def test_alias_json_existe_y_es_utf8_legible():
    if not ALIAS.is_file():
        pytest.skip("Falta alias.json: correr python sembrar_alias.py")
    datos = json.loads(ALIAS.read_text(encoding="utf-8"))

    assert len(datos) > 100
    assert all(isinstance(k, str) and isinstance(v, str) for k, v in datos.items())


def test_las_claves_de_alias_json_ya_estan_normalizadas():
    """Si se guardaran los nombres crudos, la columna saldría vacía en silencio."""
    if not ALIAS.is_file():
        pytest.skip("Falta alias.json")
    datos = json.loads(ALIAS.read_text(encoding="utf-8"))

    sin_normalizar = [k for k in datos if clave_comparacion(k) != k]
    assert not sin_normalizar, f"claves crudas en alias.json: {sin_normalizar[:5]}"


def test_cargar_alias_lee_el_archivo_generado():
    if not ALIAS.is_file():
        pytest.skip("Falta alias.json")
    alias = cargar_alias(ALIAS)

    assert corto_de("Grupo Diagnostico S.A. Dimed S.A.", alias) == "Dimed"
    assert corto_de("EMPRESA QUE NO EXISTE S.A.", alias) == ""


def test_ningun_nombre_corto_lleva_espacios_ni_saltos():
    """Va a una celda de Excel y se usa para filtrar; tiene que estar limpio."""
    if not ALIAS.is_file():
        pytest.skip("Falta alias.json")
    datos = json.loads(ALIAS.read_text(encoding="utf-8"))

    sucios = [v for v in datos.values() if v != v.strip() or "\n" in v]
    assert not sucios
