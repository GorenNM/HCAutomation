"""Escritura del Excel de salida y expansión de un expediente en varias filas.

Dos formatos, elegidos con `config.LAYOUT`:

- **`"poc"`** (el vigente): una fila por motivo de negación. `MOTIVO 1 Negación`
  pasa a llamarse `MOTIVO Negación`, `MOTIVO 2 Negación` desaparece y al final se
  añaden `Motivo #`, `Motivos totales` y `Observaciones`. 20 columnas.
- **`"clasico"`**: las 18 columnas exactas del archivo de referencia, una fila por
  expediente y los motivos repartidos en `MOTIVO 1` / `MOTIVO 2`.

El layout vive aquí entero para que revertir la POC sea cambiar una constante en
`config.py`, no reescribir el módulo.

El número de expediente se escribe como texto con hyperlink de verdad, no como
fórmula `=HYPERLINK(...)`. La referencia usa la fórmula porque así la exporta
SIPI; una fórmula sin valor cacheado se lee como `None` desde cualquier script,
y esta salida está pensada para volver a leerse.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from app.config import LAYOUT, LIMITE_CELDA_EXCEL, dir_salida
from app.models import ExtractedData, OutputRecord, SourceRow
from app.utils.rutas import base_dir
from app.utils.text import clave_comparacion, forzar_https, sanear_para_excel

log = logging.getLogger(__name__)

FILA_BANNER = 1
FILA_CABECERAS = 2
PRIMERA_FILA_DATOS = 3

# Los banners combinados de la fila 1, tal como están en el archivo de
# referencia: 'a clase 5' es literal del original, no se calcula.
_BANNERS = (("D", "H", "OPOSITOR 1 a clase 5"), ("I", "L", "OPOSITOR 2 a clase 5"))

_COMUNES_IZQUIERDA = (
    "Número de Expediente",
    "Marca",
    "Naturaleza",
    "Presenta Oposición",
    "Opositor 1",
    "Nombre corto OPOSITOR 1",
    "Art OP 1",
    "Fundada OP 1",
    "Opositor 2",
    "Nombre corto OPOSITOR 2",
    "Art Opositor 2",
    "Fundada OP 2",
)

_COMUNES_DERECHA = (
    "Apelación a la negación",
    "Titular",
    "NIZA",
)

# La descripción de productos va SIEMPRE de última (pedido del usuario,
# 2026-08-07): son párrafos larguísimos y en medio de la tabla estorbaban.
# En el layout clásico ya era la última columna del archivo de referencia.
_ULTIMA = "Descripción de Productos y Servicios"

CABECERAS = {
    "poc": (
        *_COMUNES_IZQUIERDA,
        "MOTIVO Negación",
        *_COMUNES_DERECHA,
        "Motivo #",
        "Motivos totales",
        "Observaciones",
        _ULTIMA,
    ),
    "clasico": (
        *_COMUNES_IZQUIERDA,
        "MOTIVO 1 Negación",
        "MOTIVO 2 Negación",
        *_COMUNES_DERECHA,
        _ULTIMA,
    ),
}

# Columnas que vienen del reporte de entrada. Todas las demás las produce este
# programa (de la resolución, del diccionario de alias o calculadas) y llevan
# la cabecera en amarillo para distinguirlas de un vistazo.
_DEL_REPORTE = {
    "Número de Expediente",
    "Marca",
    "Titular",
    "NIZA",
    _ULTIMA,
}

# Ancho de columna en caracteres. Solo las que se leen de un vistazo; la
# descripción de productos se deja estrecha a propósito porque son párrafos.
_ANCHOS = {
    "Número de Expediente": 18,
    "Marca": 22,
    "Naturaleza": 12,
    "Presenta Oposición": 10,
    "Opositor 1": 32,
    "Opositor 2": 32,
    "Titular": 40,
    "Descripción de Productos y Servicios": 50,
    "Observaciones": 45,
}


class ErrorEscritura(Exception):
    """No se pudo escribir el archivo. El mensaje va directo al usuario."""


# --- Expansión de filas ------------------------------------------------------


def expandir(
    source: SourceRow, extracted: ExtractedData, layout: str = LAYOUT
) -> list[OutputRecord]:
    """Un expediente -> una fila por motivo de negación.

    Sin motivos detectados sale igualmente **una** fila con la celda vacía: un
    expediente nunca desaparece de la salida por no haberse podido extraer.
    En layout `"clasico"` no hay expansión — una fila por expediente y los
    motivos van a `MOTIVO 1` / `MOTIVO 2`.
    """
    if layout == "clasico":
        return [OutputRecord(source, extracted, None, 1, len(extracted.motivos))]

    motivos: list[str | None] = list(extracted.motivos) or [None]
    return [
        OutputRecord(source, extracted, motivo, indice, len(motivos))
        for indice, motivo in enumerate(motivos, start=1)
    ]


# --- Nombres cortos ----------------------------------------------------------


def normalizar_alias(crudo: dict[str, str]) -> dict[str, str]:
    """Reindexa `{nombre tal cual: alias}` por clave de comparación.

    Sin esto, el mismo opositor escrito con o sin tilde no empata. Todo lo que
    entregue un diccionario de alias al writer debe pasar por aquí primero.
    """
    return {clave_comparacion(nombre): corto for nombre, corto in crudo.items()}


def cargar_alias(ruta: Path | None = None) -> dict[str, str]:
    """Lee `alias.json` y lo devuelve ya normalizado.

    'RED BULL GMBH' -> 'RedBull' no se deriva por ninguna regla: es criterio
    humano. El archivo se siembra con los alias que ya existen en el Excel de
    referencia; si no está, la columna sale vacía y no pasa nada más.
    """
    ruta = ruta or base_dir() / "alias.json"
    try:
        crudo = json.loads(ruta.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return {}
    except (OSError, ValueError) as error:
        log.warning("No se pudo leer %s: %s", ruta, error)
        return {}
    return normalizar_alias(crudo)


def _nombre_corto(nombre: str, alias: dict[str, str]) -> str:
    return alias.get(clave_comparacion(nombre), "") if nombre else ""


# --- Valores de una fila -----------------------------------------------------


def _opositor(registro: OutputRecord, indice: int, alias: dict[str, str]) -> list[str]:
    """Las cuatro celdas de un opositor. Vacías si ese opositor no existe."""
    opositores = registro.extracted.opositores
    if indice >= len(opositores):
        return ["", "", "", ""]
    opositor = opositores[indice]
    return [
        opositor.nombre,
        _nombre_corto(opositor.nombre, alias),
        ", ".join(opositor.articulos),
        opositor.fundada or "",
    ]


def _apelacion(registro: OutputRecord) -> str:
    """'SI' / 'no' / '' — los tres valores que usa el archivo de referencia."""
    apelacion = registro.extracted.apelacion
    if apelacion is None:
        return ""
    return "SI" if apelacion else "no"


def valores(registro: OutputRecord, alias: dict[str, str], layout: str) -> list[str]:
    """Fila completa, en el orden de `CABECERAS[layout]`, ya saneada.

    `alias` tiene que venir de `cargar_alias` o `normalizar_alias`: sus claves
    son claves de comparación, no nombres tal cual.
    """
    extracted = registro.extracted
    motivos = extracted.motivos

    if layout == "clasico":
        celdas_motivo = [
            motivos[0] if motivos else "",
            motivos[1] if len(motivos) > 1 else "",
        ]
        extra: list[str] = []
    else:
        celdas_motivo = [registro.motivo or ""]
        extra = [
            str(registro.motivo_indice),
            str(registro.motivos_total),
            "; ".join(extracted.avisos),
        ]

    fila = [
        registro.source.expediente,
        registro.source.marca,
        extracted.naturaleza or "",
        "Sí" if extracted.presenta_oposicion else "No",
        *_opositor(registro, 0, alias),
        *_opositor(registro, 1, alias),
        *celdas_motivo,
        _apelacion(registro),
        registro.source.titular,
        registro.source.niza,
        *extra,
        registro.source.descripcion,
    ]
    return [sanear_para_excel(str(celda), LIMITE_CELDA_EXCEL) for celda in fila]


# --- Escritura ---------------------------------------------------------------


# Bordes: negro fino en los datos y más grueso en cabeceras y banners.
# "medium" y no "thick" a propósito: con 20 columnas, el trazo más gordo de
# openpyxl ensucia más de lo que separa.
_LADO_FINO = Side(style="thin", color="000000")
_LADO_GRUESO = Side(style="medium", color="000000")
BORDE_DATOS = Border(
    left=_LADO_FINO, right=_LADO_FINO, top=_LADO_FINO, bottom=_LADO_FINO
)
BORDE_TITULO = Border(
    left=_LADO_GRUESO, right=_LADO_GRUESO, top=_LADO_GRUESO, bottom=_LADO_GRUESO
)

# Azul y subrayado de hipervínculo de Excel. La celda ya lleva un hyperlink de
# verdad, pero con la fuente por defecto no PARECE clicable y nadie hace clic.
FUENTE_ENLACE = Font(color="0563C1", underline="single")

# Morado para las filas sin opositor: la mitad de las columnas (D–L, los dos
# bloques de opositor) sale vacía y no se distingue de un fallo de extracción.
# El color dice «aquí no hay nada que poner», que es distinto de «no se pudo».
FONDO_SIN_OPOSITOR = PatternFill("solid", fgColor="917AC3")


def _escribir_encabezado(hoja, cabeceras: tuple[str, ...]) -> None:
    for izquierda, derecha, texto in _BANNERS:
        hoja.merge_cells(f"{izquierda}{FILA_BANNER}:{derecha}{FILA_BANNER}")
        celda = hoja[f"{izquierda}{FILA_BANNER}"]
        celda.value = texto
        celda.alignment = Alignment(horizontal="center")
        celda.font = Font(bold=True)
        # El borde de un rango combinado solo se pinta entero si lo llevan
        # todas las celdas del rango, no solo la primera.
        for columna in range(
            column_index_from_string(izquierda), column_index_from_string(derecha) + 1
        ):
            hoja.cell(row=FILA_BANNER, column=columna).border = BORDE_TITULO

    heredado = PatternFill("solid", fgColor="DDDDDD")
    generado = PatternFill("solid", fgColor="FFFF00")
    for columna, cabecera in enumerate(cabeceras, start=1):
        celda = hoja.cell(row=FILA_CABECERAS, column=columna, value=cabecera)
        celda.font = Font(bold=True)
        celda.fill = heredado if cabecera in _DEL_REPORTE else generado
        celda.border = BORDE_TITULO
        celda.alignment = Alignment(vertical="center", wrap_text=True)
        hoja.column_dimensions[get_column_letter(columna)].width = _ANCHOS.get(
            cabecera, 14
        )

    hoja.freeze_panes = f"A{PRIMERA_FILA_DATOS}"


def ruta_por_defecto(carpeta: Path | None = None, momento: str = "") -> Path:
    """`salida\\Negacion_marcas_20260806_1435.xlsx`.

    Lleva marca de tiempo para que dos corridas no se pisen, que es lo que pasa
    en cuanto alguien repite el proceso sin renombrar nada.
    """
    marca = momento or datetime.now().strftime("%Y%m%d_%H%M%S")
    return (carpeta or dir_salida()) / f"Negacion_marcas_{marca}.xlsx"


def escribir(
    registros: list[OutputRecord],
    ruta: Path,
    layout: str = LAYOUT,
    alias: dict[str, str] | None = None,
) -> Path:
    """Vuelca los registros a un `.xlsx`. Devuelve la ruta escrita.

    Con cero registros escribe igualmente el archivo, solo con las cabeceras:
    un archivo vacío pero abrible dice mucho más que ningún archivo.
    """
    if layout not in CABECERAS:
        raise ErrorEscritura(
            f"Layout desconocido: {layout!r}. Valores válidos: "
            f"{', '.join(sorted(CABECERAS))}."
        )

    cabeceras = CABECERAS[layout]
    alias = cargar_alias() if alias is None else alias

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Hoja1"
    _escribir_encabezado(hoja, cabeceras)

    for numero, registro in enumerate(registros, start=PRIMERA_FILA_DATOS):
        fila = valores(registro, alias, layout)
        sin_opositor = not registro.extracted.opositores
        for columna, valor in enumerate(fila, start=1):
            # El `data_type = "s"` no sobra: una marca que empiece por '=' la
            # guarda openpyxl como FÓRMULA y Excel la evalúa al abrir el
            # archivo. Aquí todo es texto, siempre.
            celda = hoja.cell(row=numero, column=columna, value=valor)
            celda.data_type = "s"
            celda.border = BORDE_DATOS
            if sin_opositor:
                celda.fill = FONDO_SIN_OPOSITOR
        enlace = registro.source.case_url
        if enlace:
            celda = hoja.cell(row=numero, column=1)
            # El reporte exporta http:// y el puerto 80 de SIPI no
            # responde: un enlace http se queda colgado hasta el timeout
            # del navegador. Al Excel de salida va siempre https.
            celda.hyperlink = forzar_https(enlace)
            celda.font = FUENTE_ENLACE

    ruta = Path(ruta)
    try:
        ruta.parent.mkdir(parents=True, exist_ok=True)
        libro.save(ruta)
    except PermissionError:
        # El caso real: el usuario tiene el archivo abierto en Excel.
        raise ErrorEscritura(
            f"No se pudo escribir «{ruta.name}»: el archivo está abierto en Excel "
            f"o no hay permiso sobre «{ruta.parent}». Ciérrelo e intente de nuevo."
        ) from None
    except OSError as error:
        raise ErrorEscritura(f"No se pudo escribir «{ruta}»: {error}") from error
    finally:
        libro.close()

    log.info("Escritos %d registros en %s", len(registros), ruta)
    return ruta
