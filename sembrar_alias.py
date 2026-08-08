"""Genera `alias.json` a partir del archivo de referencia.

    python sembrar_alias.py "Negacion marcas con información extra.xlsx"

El «Nombre corto» de un opositor no se deduce por ninguna regla: `RED BULL GMBH`
→ `RedBull` es criterio humano. Lo único que se puede hacer es reaprovechar los
que alguien ya escribió a mano, y eso es este script.

El archivo resultante se deja junto al `.exe`. Si no está, la columna sale vacía
y no pasa nada más: `writer.cargar_alias()` devuelve un diccionario vacío.

Las claves NO son el nombre tal cual sino la clave de comparación
(`utils.text.clave_comparacion`), que ignora tildes, mayúsculas, puntuación y
sufijos societarios. Así `Grupo Diagnóstico S.A. Dimed S.A.` empata con
`GRUPO DIAGNOSTICO DIMED`, que es como aparece en algunas resoluciones.
"""

from __future__ import annotations

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from app.utils.text import clave_comparacion, normalizar

# Columnas del archivo de referencia (base 0): pares nombre / nombre corto.
PARES_DE_COLUMNAS = ((4, 5), (8, 9))
PRIMERA_FILA = 3


def leer_pares(ruta: Path) -> list[tuple[str, str]]:
    libro = openpyxl.load_workbook(ruta, read_only=True)
    try:
        hoja = libro[libro.sheetnames[0]]
        pares = []
        for fila in hoja.iter_rows(min_row=PRIMERA_FILA, values_only=True):
            for col_nombre, col_corto in PARES_DE_COLUMNAS:
                if len(fila) <= col_corto:
                    continue
                nombre, corto = fila[col_nombre], fila[col_corto]
                if not nombre or not corto:
                    continue
                # El original trae saltos de línea dentro de algunos nombres.
                nombre, corto = normalizar(str(nombre)), str(corto).strip()
                if nombre and corto:
                    pares.append((nombre, corto))
        return pares
    finally:
        libro.close()


def construir(pares: list[tuple[str, str]]) -> tuple[dict[str, str], list[str]]:
    """Devuelve (alias por clave de comparación, avisos de conflicto).

    Cuando el mismo opositor tiene dos nombres cortos distintos —pasa: hay
    erratas en el archivo de referencia— gana el más frecuente, y se avisa para
    que un humano pueda corregirlo a mano. En empate gana el primero que
    aparece en el archivo, que es arbitrario pero determinista: dos corridas
    sobre el mismo Excel dan el mismo `alias.json`.
    """
    votos: dict[str, Counter[str]] = defaultdict(Counter)
    escrituras: dict[str, set[str]] = defaultdict(set)
    for nombre, corto in pares:
        clave = clave_comparacion(nombre)
        if not clave:
            continue
        votos[clave][corto] += 1
        escrituras[clave].add(nombre)

    alias: dict[str, str] = {}
    avisos: list[str] = []
    for clave, cuenta in sorted(votos.items()):
        elegido, veces = cuenta.most_common(1)[0]
        alias[clave] = elegido
        if len(cuenta) > 1:
            resto = ", ".join(
                f"{corto} ({n})" for corto, n in cuenta.most_common()[1:]
            )
            avisos.append(
                f"«{sorted(escrituras[clave])[0]}» tiene varios nombres cortos: "
                f"se usa {elegido} ({veces}); se descartan {resto}"
            )
    return alias, avisos


def main(argv: list[str]) -> int:
    origen = Path(argv[1]) if len(argv) > 1 else Path(
        "Negacion marcas con información extra.xlsx"
    )
    if not origen.is_file():
        print(f"No existe «{origen}».", file=sys.stderr)
        return 1

    pares = leer_pares(origen)
    alias, avisos = construir(pares)

    destino = Path("alias.json")
    destino.write_text(
        json.dumps(alias, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )

    print(f"{len(pares)} menciones -> {len(alias)} opositores en {destino}")
    for aviso in avisos:
        print(f"  CONFLICTO  {aviso}")
    if avisos:
        print(
            "\nSon erratas del archivo de referencia. Si alguna elección no "
            "convence, se edita alias.json a mano: no se vuelve a generar solo."
        )
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
