"""Pruebas de la escritura del Excel de salida.

Las aserciones no se ablandan: si algo pasa por aquí es porque el writer
aguanta, no porque la prueba se haya conformado.
"""

from __future__ import annotations

import json
import os
import sys

import openpyxl
import pytest

from app.config import LIMITE_CELDA_EXCEL
from app.excel.writer import (
    CABECERAS,
    FILA_CABECERAS,
    PRIMERA_FILA_DATOS,
    ErrorEscritura,
    cargar_alias,
    escribir,
    expandir,
    normalizar_alias,
    ruta_por_defecto,
    valores,
)
from app.models import ExtractedData, Opositor, OutputRecord, SourceRow

# Las 18 cabeceras del archivo de referencia, copiadas literalmente de él
# (sin los espacios sobrantes que trae el original en algunas celdas).
REFERENCIA_18 = (
    "Número de Expediente",
    "Marca",
    "Naturaleza",
    "Presenta Oposición",
    "Opositor 1",
    "Nombre corto OPOSITOR 1",
    "Art OP 1",
    "Fundada OP 1",
    "Opositor 2",
    "Nombre corto OPOSITOR 2",
    "Art Opositor 2",
    "Fundada OP 2",
    "MOTIVO 1 Negación",
    "MOTIVO 2 Negación",
    "Apelación a la negación",
    "Titular",
    "NIZA",
    "Descripción de Productos y Servicios",
)


def origen(expediente: str = "SD2022/0000017", **kwargs) -> SourceRow:
    datos = {
        "expediente": expediente,
        "case_url": "https://sipi.sic.gov.co/sipi/View.ashx?3857028",
        "marca": "MALTAVITAN",
        "titular": "Pieter Carl Alexander Heshusius Florez",
        "niza": "5",
        "descripcion": "5. Productos farmacéuticos.",
        "bajo_oposicion": "No",
        "fila": 12,
    }
    datos.update(kwargs)
    return SourceRow(**datos)


def extraido(**kwargs) -> ExtractedData:
    return ExtractedData(**kwargs)


def registros_de(source: SourceRow, data: ExtractedData, layout="poc"):
    return expandir(source, data, layout)


def releer(ruta):
    libro = openpyxl.load_workbook(ruta)
    hoja = libro["Hoja1"]
    cabeceras = [c.value for c in hoja[FILA_CABECERAS]]
    filas = [
        [celda.value for celda in fila]
        for fila in hoja.iter_rows(min_row=PRIMERA_FILA_DATOS)
    ]
    return hoja, cabeceras, filas


# --- Expansión de filas ------------------------------------------------------


@pytest.mark.parametrize(
    "motivos,filas_esperadas",
    [([], 1), (["136a"], 1), (["136a", "136h"], 2), (["1", "2", "3", "4", "5"], 5)],
)
def test_una_fila_por_motivo(motivos, filas_esperadas):
    filas = expandir(origen(), extraido(motivos=motivos), "poc")
    assert len(filas) == filas_esperadas


def test_sin_motivos_el_expediente_no_desaparece():
    """Requisito explícito: nunca se pierde un expediente por no extraer nada."""
    (fila,) = expandir(origen(), extraido(motivos=[]), "poc")

    assert fila.motivo is None
    assert fila.motivo_indice == 1
    assert fila.motivos_total == 1


def test_los_indices_van_en_orden_de_aparicion():
    filas = expandir(origen(), extraido(motivos=["136h", "136a", "135b"]), "poc")

    assert [f.motivo for f in filas] == ["136h", "136a", "135b"]
    assert [f.motivo_indice for f in filas] == [1, 2, 3]
    assert {f.motivos_total for f in filas} == {3}


def test_el_resto_de_la_informacion_se_repite_en_cada_fila():
    fuente = origen()
    datos = extraido(motivos=["136a", "136h"], naturaleza="Mixta")
    filas = expandir(fuente, datos, "poc")

    assert all(f.source is fuente for f in filas)
    assert all(f.extracted is datos for f in filas)


def test_el_layout_clasico_no_expande():
    (fila,) = expandir(origen(), extraido(motivos=["136a", "136h", "136b"]), "clasico")
    assert fila.motivos_total == 3


# --- Caso 35: salida de 0 registros ------------------------------------------


def test_caso_35_cero_registros_da_un_archivo_valido(tmp_path):
    ruta = escribir([], tmp_path / "vacio.xlsx", layout="poc", alias={})

    hoja, cabeceras, filas = releer(ruta)
    assert cabeceras == list(CABECERAS["poc"])
    assert filas == []
    assert hoja.max_row == FILA_CABECERAS


# --- Caso 36: caracteres ilegales para xlsx ----------------------------------


def test_caso_36_los_caracteres_de_control_no_revientan_openpyxl(tmp_path):
    """openpyxl lanza excepción con \\x00 y compañía; hay que sanear antes."""
    veneno = "MARCA\x00CON\x01BASURA\x0bDE\x1fCONTROL"
    fuente = origen(marca=veneno, titular="titular\x07campana")
    datos = extraido(motivos=["136a"], avisos=["aviso\x00roto"])

    ruta = escribir(
        registros_de(fuente, datos), tmp_path / "sucio.xlsx", "poc", alias={}
    )

    _, _, filas = releer(ruta)
    marca = filas[0][1]
    assert marca == "MARCACONBASURADECONTROL"
    assert not any(ord(c) < 32 for c in marca)


def test_caso_36_los_saltos_de_linea_reales_sobreviven(tmp_path):
    """\\n y \\t sí son legales en una celda: sanear no puede comérselos."""
    fuente = origen(descripcion="5. Productos.\nOtra línea.\tTabulada.")

    ruta = escribir(
        registros_de(fuente, extraido(motivos=["136a"])),
        tmp_path / "saltos.xlsx",
        "poc",
        alias={},
    )

    _, cabeceras, filas = releer(ruta)
    columna = cabeceras.index("Descripción de Productos y Servicios")
    assert filas[0][columna] == "5. Productos.\nOtra línea.\tTabulada."


@pytest.mark.parametrize(
    "texto", ["=1+1", "=SUM(A1:A9)", "+34", "-CALC()", "@SUM(1)", '=cmd|"/c calc"!A1']
)
def test_caso_36_un_texto_que_empieza_por_igual_no_se_vuelve_formula(tmp_path, texto):
    """Inyección de fórmulas: una marca `=1+1` la evaluaría Excel al abrir."""
    ruta = escribir(
        registros_de(origen(marca=texto), extraido(motivos=["136a"])),
        tmp_path / "formula.xlsx",
        "poc",
        alias={},
    )

    celda = openpyxl.load_workbook(ruta)["Hoja1"].cell(row=PRIMERA_FILA_DATOS, column=2)
    assert celda.data_type == "s", "quedó guardada como fórmula"
    assert celda.value == texto


def test_caso_36_un_subrogado_suelto_no_deja_el_archivo_ilegible(tmp_path):
    """openpyxl guarda \\ud800 sin protestar y luego el .xlsx no se puede abrir."""
    fuente = origen(titular="mal\ud800texto", marca="marca\udfffrota")

    ruta = escribir(
        registros_de(fuente, extraido(motivos=["136a"])),
        tmp_path / "subrogados.xlsx",
        "poc",
        alias={},
    )

    _, cabeceras, filas = releer(ruta)  # aquí es donde reventaba
    assert filas[0][cabeceras.index("Titular")] == "maltexto"
    assert filas[0][cabeceras.index("Marca")] == "marcarota"


# --- Caso 37: celda por encima del límite de Excel ---------------------------


def test_caso_37_una_descripcion_gigante_se_trunca_con_marca(tmp_path):
    fuente = origen(descripcion="x" * (LIMITE_CELDA_EXCEL + 5_000))

    ruta = escribir(
        registros_de(fuente, extraido(motivos=["136a"])),
        tmp_path / "gigante.xlsx",
        "poc",
        alias={},
    )

    _, cabeceras, filas = releer(ruta)
    celda = filas[0][cabeceras.index("Descripción de Productos y Servicios")]
    assert len(celda) == LIMITE_CELDA_EXCEL
    assert celda.endswith("[truncado]")


def test_caso_37_ninguna_celda_supera_nunca_el_limite(tmp_path):
    """Todos los campos, no solo la descripción: el titular también viene largo."""
    enorme = "y" * (LIMITE_CELDA_EXCEL * 2)
    fuente = origen(marca=enorme, titular=enorme, niza=enorme, descripcion=enorme)
    datos = extraido(
        motivos=["136a"],
        naturaleza=enorme,
        avisos=[enorme],
        opositores=[Opositor(nombre=enorme, articulos=[enorme], fundada="SI")],
    )

    ruta = escribir(
        registros_de(fuente, datos), tmp_path / "todo_enorme.xlsx", "poc", alias={}
    )

    _, _, filas = releer(ruta)
    assert all(len(celda) <= LIMITE_CELDA_EXCEL for celda in filas[0] if celda)


# --- Caso 38: ida y vuelta ---------------------------------------------------


def test_caso_38_lo_que_se_escribe_es_lo_que_se_relee(tmp_path):
    fuente = origen()
    datos = extraido(
        naturaleza="Mixta",
        presenta_oposicion=True,
        apelacion=True,
        motivos=["136a", "136h"],
        avisos=["se citaron 3 opositores"],
        opositores=[
            Opositor("NESTLE SA", ["136a", "136h"], "SI"),
            Opositor("KRAFT FOODS", ["136a"], "NO"),
        ],
    )
    registros = registros_de(fuente, datos)

    ruta = escribir(registros, tmp_path / "vuelta.xlsx", "poc", alias={})

    _, cabeceras, filas = releer(ruta)
    assert len(filas) == 2
    for fila_leida, registro in zip(filas, registros):
        esperado = valores(registro, {}, "poc")
        assert [celda or "" for celda in fila_leida] == esperado


def test_caso_38_el_hyperlink_del_expediente_sobrevive(tmp_path):
    fuente = origen()
    ruta = escribir(
        registros_de(fuente, extraido(motivos=["136a"])),
        tmp_path / "enlace.xlsx",
        "poc",
        alias={},
    )

    hoja = openpyxl.load_workbook(ruta)["Hoja1"]
    celda = hoja.cell(row=PRIMERA_FILA_DATOS, column=1)

    assert celda.value == "SD2022/0000017"  # texto, no fórmula
    assert celda.hyperlink is not None
    assert celda.hyperlink.target == fuente.case_url


def test_un_expediente_sin_url_no_revienta_al_escribir(tmp_path):
    ruta = escribir(
        registros_de(origen(case_url=""), extraido(motivos=["136a"])),
        tmp_path / "sin_url.xlsx",
        "poc",
        alias={},
    )

    celda = openpyxl.load_workbook(ruta)["Hoja1"].cell(
        row=PRIMERA_FILA_DATOS, column=1
    )
    assert celda.hyperlink is None


# --- Los dos layouts ---------------------------------------------------------


def test_el_layout_clasico_reproduce_las_18_columnas_de_la_referencia():
    assert CABECERAS["clasico"] == REFERENCIA_18


def test_el_layout_poc_renombra_motivo_1_y_elimina_motivo_2():
    poc = CABECERAS["poc"]

    assert "MOTIVO Negación" in poc
    assert "MOTIVO 1 Negación" not in poc
    assert "MOTIVO 2 Negación" not in poc
    # La descripción de productos cierra la tabla (pedido del usuario,
    # 2026-08-07): es un párrafo largo y en medio estorbaba.
    assert poc[-4:] == (
        "Motivo #",
        "Motivos totales",
        "Observaciones",
        "Descripción de Productos y Servicios",
    )
    assert len(poc) == 20


def test_el_clasico_reparte_los_motivos_en_dos_columnas(tmp_path):
    datos = extraido(motivos=["136a", "136h"])
    ruta = escribir(
        expandir(origen(), datos, "clasico"),
        tmp_path / "clasico.xlsx",
        "clasico",
        alias={},
    )

    _, cabeceras, filas = releer(ruta)
    assert len(filas) == 1
    assert filas[0][cabeceras.index("MOTIVO 1 Negación")] == "136a"
    assert filas[0][cabeceras.index("MOTIVO 2 Negación")] == "136h"


def test_el_clasico_pierde_el_tercer_motivo_a_proposito():
    """Es la limitación conocida del formato viejo, no un bug del writer."""
    datos = extraido(motivos=["136a", "136h", "135b"])
    fila = valores(expandir(origen(), datos, "clasico")[0], {}, "clasico")

    assert "135b" not in fila


def test_un_layout_desconocido_falla_con_mensaje_util(tmp_path):
    with pytest.raises(ErrorEscritura) as error:
        escribir([], tmp_path / "x.xlsx", layout="inventado", alias={})

    assert "inventado" in str(error.value)
    assert "poc" in str(error.value) and "clasico" in str(error.value)


# --- Opositores --------------------------------------------------------------


def test_sin_opositores_las_ocho_celdas_quedan_vacias():
    fila = valores(expandir(origen(), extraido(motivos=["136a"]))[0], {}, "poc")
    assert fila[4:12] == [""] * 8


def test_con_un_solo_opositor_el_segundo_bloque_queda_vacio():
    datos = extraido(motivos=["136a"], opositores=[Opositor("NESTLE SA", ["136a"], "SI")])
    fila = valores(expandir(origen(), datos)[0], {}, "poc")

    assert fila[4] == "NESTLE SA"
    assert fila[8:12] == [""] * 4


def test_un_tercer_opositor_no_agrega_columnas():
    """La salida se queda con 2 opositores; el resto va a Observaciones."""
    datos = extraido(
        motivos=["136a"],
        opositores=[Opositor(f"OPOSITOR {n}") for n in range(1, 6)],
        avisos=["se detectaron 5 opositores, solo se reportan 2"],
    )
    fila = valores(expandir(origen(), datos)[0], {}, "poc")

    assert fila[4] == "OPOSITOR 1"
    assert fila[8] == "OPOSITOR 2"
    assert "OPOSITOR 3" not in fila
    assert "5 opositores" in fila[list(CABECERAS["poc"]).index("Observaciones")]


def test_los_articulos_de_un_opositor_van_separados_por_coma():
    datos = extraido(
        motivos=["136a"],
        opositores=[Opositor("X", ["136a", "136b", "136h"], "SI")],
    )
    assert valores(expandir(origen(), datos)[0], {}, "poc")[6] == "136a, 136b, 136h"


@pytest.mark.parametrize(
    "apelacion,esperado", [(True, "SI"), (False, "no"), (None, "")]
)
def test_la_apelacion_usa_los_tres_valores_de_la_referencia(apelacion, esperado):
    datos = extraido(motivos=["136a"], apelacion=apelacion)
    fila = valores(expandir(origen(), datos)[0], {}, "poc")
    assert fila[13] == esperado


def test_los_campos_vacios_no_escriben_la_palabra_none(tmp_path):
    """`str(None)` en una celda es el error tonto que ensucia todo el reporte."""
    ruta = escribir(
        registros_de(origen(), extraido()), tmp_path / "nones.xlsx", "poc", alias={}
    )

    _, _, filas = releer(ruta)
    assert "None" not in [celda for celda in filas[0] if celda]


# --- Nombres cortos (alias) --------------------------------------------------


def test_sin_archivo_de_alias_la_columna_sale_vacia(tmp_path):
    assert cargar_alias(tmp_path / "no_existe.json") == {}


def test_un_alias_json_corrupto_no_tumba_la_corrida(tmp_path):
    roto = tmp_path / "alias.json"
    roto.write_text("{esto no es json", encoding="utf-8")

    assert cargar_alias(roto) == {}


def test_el_alias_empata_aunque_cambien_tildes_mayusculas_y_sufijos(tmp_path):
    ruta = tmp_path / "alias.json"
    ruta.write_text(
        json.dumps({"Grupo Diagnóstico S.A. Dimed S.A.": "Dimed"}), encoding="utf-8"
    )
    alias = cargar_alias(ruta)

    datos = extraido(
        motivos=["136a"], opositores=[Opositor("GRUPO DIAGNOSTICO DIMED", ["136a"], "SI")]
    )
    assert valores(expandir(origen(), datos)[0], alias, "poc")[5] == "Dimed"


def test_la_enie_no_funde_dos_opositores_distintos(tmp_path):
    """`MUÑOZ` y `MUNOZ` son empresas distintas; el alias no puede confundirlas."""
    ruta = tmp_path / "alias.json"
    ruta.write_text(json.dumps({"MUÑOZ S.A.": "Muñoz"}), encoding="utf-8")
    alias = cargar_alias(ruta)

    datos = extraido(motivos=["136a"], opositores=[Opositor("MUNOZ S.A.")])
    assert valores(expandir(origen(), datos)[0], alias, "poc")[5] == ""


def test_un_alias_sin_normalizar_no_empata_y_por_eso_existe_normalizar_alias():
    """Trampa real: pasar el diccionario crudo deja la columna vacía en silencio."""
    crudo = {"Grupo Diagnóstico S.A. Dimed S.A.": "Dimed"}
    datos = extraido(
        motivos=["136a"], opositores=[Opositor("Grupo Diagnostico S.A. Dimed S.A.")]
    )
    registro = expandir(origen(), datos)[0]

    assert valores(registro, crudo, "poc")[5] == ""
    assert valores(registro, normalizar_alias(crudo), "poc")[5] == "Dimed"


def test_un_opositor_sin_alias_deja_la_celda_vacia_sin_inventar():
    datos = extraido(motivos=["136a"], opositores=[Opositor("EMPRESA DESCONOCIDA")])
    assert valores(expandir(origen(), datos)[0], {"otra": "Otra"}, "poc")[5] == ""


# --- Escritura: rutas y fallos -----------------------------------------------


def test_la_carpeta_de_salida_se_crea_si_no_existe(tmp_path):
    ruta = tmp_path / "no" / "existe" / "aun" / "salida.xlsx"
    assert escribir([], ruta, "poc", alias={}).is_file()


def test_dos_corridas_seguidas_no_se_pisan():
    primera = ruta_por_defecto(momento="20260806_143000")
    segunda = ruta_por_defecto(momento="20260806_143001")
    assert primera != segunda
    assert primera.suffix == ".xlsx"


def test_una_ruta_con_acentos_y_espacios_funciona(tmp_path):
    """Caso 42: `C:\\Users\\José\\Mis documentos\\`."""
    ruta = tmp_path / "José Ramírez" / "Mis documentos" / "salida ñandú.xlsx"

    escrita = escribir(
        registros_de(origen(), extraido(motivos=["136a"])), ruta, "poc", alias={}
    )

    _, _, filas = releer(escrita)
    assert filas[0][0] == "SD2022/0000017"


@pytest.mark.skipif(
    sys.platform == "win32" or os.geteuid() == 0,
    reason="chmod no bloquea a root ni funciona igual en Windows",
)
def test_sin_permiso_de_escritura_el_mensaje_es_para_un_humano(tmp_path):
    """Caso 39: nada de `PermissionError` crudo en pantalla."""
    carpeta = tmp_path / "bloqueada"
    carpeta.mkdir()
    carpeta.chmod(0o500)
    try:
        with pytest.raises(ErrorEscritura) as error:
            escribir([], carpeta / "salida.xlsx", "poc", alias={})
    finally:
        carpeta.chmod(0o700)

    mensaje = str(error.value)
    assert "salida.xlsx" in mensaje
    assert "Excel" in mensaje and "intente de nuevo" in mensaje
    assert "PermissionError" not in mensaje
    assert "Traceback" not in mensaje


def test_una_ruta_imposible_da_error_de_escritura_no_una_traza(tmp_path):
    """La carpeta contenedora ya existe, pero es un archivo."""
    estorbo = tmp_path / "no_soy_carpeta"
    estorbo.write_text("ocupado", encoding="utf-8")

    with pytest.raises(ErrorEscritura) as error:
        escribir([], estorbo / "salida.xlsx", "poc", alias={})

    assert "No se pudo escribir" in str(error.value)


def test_el_archivo_existente_se_sobreescribe(tmp_path):
    ruta = tmp_path / "salida.xlsx"
    escribir(registros_de(origen(), extraido(motivos=["a", "b"])), ruta, "poc", alias={})
    escribir(registros_de(origen(), extraido(motivos=["c"])), ruta, "poc", alias={})

    _, _, filas = releer(ruta)
    assert len(filas) == 1


# --- Encabezado --------------------------------------------------------------


def test_la_fila_banner_conserva_las_dos_celdas_combinadas(tmp_path):
    ruta = escribir([], tmp_path / "banner.xlsx", "poc", alias={})
    hoja = openpyxl.load_workbook(ruta)["Hoja1"]

    rangos = {str(r) for r in hoja.merged_cells.ranges}
    assert rangos == {"D1:H1", "I1:L1"}
    assert hoja["D1"].value == "OPOSITOR 1 a clase 5"
    assert hoja["I1"].value == "OPOSITOR 2 a clase 5"


def test_las_cabeceras_quedan_congeladas(tmp_path):
    ruta = escribir([], tmp_path / "panes.xlsx", "poc", alias={})
    assert openpyxl.load_workbook(ruta)["Hoja1"].freeze_panes == "A3"


# --- Volumen -----------------------------------------------------------------


def test_dos_mil_registros_se_escriben_y_se_releen_enteros(tmp_path):
    """El reporte real son 987 expedientes; con multimotivo pasan de 1 000 filas."""
    registros = [
        OutputRecord(
            origen(f"SD2022/{numero:07d}"),
            extraido(motivos=["136a", "136h"], naturaleza="Mixta"),
            motivo,
            indice,
            2,
        )
        for numero in range(1_000)
        for indice, motivo in enumerate(["136a", "136h"], start=1)
    ]

    ruta = escribir(registros, tmp_path / "volumen.xlsx", "poc", alias={})

    _, _, filas = releer(ruta)
    assert len(filas) == 2_000
    assert filas[0][0] == "SD2022/0000000"
    assert filas[-1][0] == "SD2022/0000999"


# --- Contra los datos reales -------------------------------------------------


def test_el_expediente_real_multimotivo_da_dos_filas_identicas_salvo_el_motivo():
    """SD2022/0097089: 136a y 136h, dos opositores. Es el caso que motivó la POC."""
    fuente = origen("SD2022/0097089", marca="MASTER")
    datos = extraido(
        naturaleza="Mixta",
        presenta_oposicion=True,
        motivos=["136a", "136h"],
        opositores=[
            Opositor("NESTLE SA", ["136a", "136h"], "SI"),
            Opositor("KRAFT FOODS", ["136a"], "SI"),
        ],
    )

    primera, segunda = [
        valores(registro, {}, "poc") for registro in expandir(fuente, datos)
    ]

    posicion = list(CABECERAS["poc"]).index("MOTIVO Negación")
    numero = list(CABECERAS["poc"]).index("Motivo #")
    assert primera[posicion] == "136a"
    assert segunda[posicion] == "136h"
    assert (primera[numero], segunda[numero]) == ("1", "2")
    del primera[posicion], segunda[posicion]
    del primera[numero - 1], segunda[numero - 1]
    # Todo lo demás se repite, salvo el motivo y su número.
    assert primera == segunda


# --- Correcciones post-entrega (2026-08-07) ----------------------------------


def test_el_enlace_sale_https_aunque_el_reporte_traiga_http(tmp_path):
    """El reporte de SIPI exporta http:// y el puerto 80 no responde: el enlace
    del Excel de salida se colgaba hasta el timeout del navegador."""
    fuente = origen(case_url="http://sipi.sic.gov.co/sipi/View.ashx?3857028")
    ruta = escribir(
        registros_de(fuente, extraido(motivos=["136a"])),
        tmp_path / "https.xlsx",
        "poc",
        alias={},
    )

    celda = openpyxl.load_workbook(ruta)["Hoja1"].cell(
        row=PRIMERA_FILA_DATOS, column=1
    )
    assert celda.hyperlink.target == "https://sipi.sic.gov.co/sipi/View.ashx?3857028"


def test_el_expediente_con_enlace_se_ve_azul_y_subrayado(tmp_path):
    """El hyperlink existía pero con la fuente por defecto no parecía clicable."""
    ruta = escribir(
        registros_de(origen(), extraido(motivos=["136a"])),
        tmp_path / "estilo.xlsx",
        "poc",
        alias={},
    )

    celda = openpyxl.load_workbook(ruta)["Hoja1"].cell(
        row=PRIMERA_FILA_DATOS, column=1
    )
    assert celda.font.underline == "single"
    assert str(celda.font.color.rgb).endswith("0563C1")


def test_cabeceras_amarillas_solo_en_las_columnas_que_calcula_el_programa(tmp_path):
    """Amarillo = dato producido por el programa; gris = venía del reporte."""
    ruta = escribir(
        registros_de(origen(), extraido(motivos=["136a"])),
        tmp_path / "colores.xlsx",
        "poc",
        alias={},
    )

    hoja = openpyxl.load_workbook(ruta)["Hoja1"]
    del_reporte = {
        "Número de Expediente",
        "Marca",
        "Titular",
        "NIZA",
        "Descripción de Productos y Servicios",
    }
    for columna, cabecera in enumerate(CABECERAS["poc"], start=1):
        color = str(hoja.cell(row=FILA_CABECERAS, column=columna).fill.fgColor.rgb)
        esperado = "DDDDDD" if cabecera in del_reporte else "FFFF00"
        assert color.endswith(esperado), f"{cabecera}: {color}"


def test_las_filas_sin_opositor_van_moradas_y_las_demas_no(tmp_path):
    """Sin opositor, media fila sale vacía; el morado dice que es lo correcto."""
    ruta_sin = escribir(
        registros_de(origen(), extraido(motivos=["136a"])),
        tmp_path / "sin.xlsx",
        "poc",
        alias={},
    )
    ruta_con = escribir(
        registros_de(
            origen(),
            extraido(motivos=["136a"], opositores=[Opositor("NESTLE SA", ["136a"], "SI")]),
        ),
        tmp_path / "con.xlsx",
        "poc",
        alias={},
    )

    columnas = len(CABECERAS["poc"])
    hoja_sin = openpyxl.load_workbook(ruta_sin)["Hoja1"]
    for columna in range(1, columnas + 1):
        celda = hoja_sin.cell(row=PRIMERA_FILA_DATOS, column=columna)
        assert str(celda.fill.fgColor.rgb).endswith("917AC3"), columna

    hoja_con = openpyxl.load_workbook(ruta_con)["Hoja1"]
    for columna in range(1, columnas + 1):
        celda = hoja_con.cell(row=PRIMERA_FILA_DATOS, column=columna)
        assert not str(celda.fill.fgColor.rgb).endswith("917AC3"), columna


def test_todas_las_celdas_con_datos_llevan_borde(tmp_path):
    """Fino en los datos, más grueso en cabeceras y banners."""
    registros = registros_de(origen(), extraido(motivos=["136a", "136h"]))
    ruta = escribir(registros, tmp_path / "bordes.xlsx", "poc", alias={})

    hoja = openpyxl.load_workbook(ruta)["Hoja1"]
    columnas = len(CABECERAS["poc"])
    for columna in range(1, columnas + 1):
        assert hoja.cell(row=FILA_CABECERAS, column=columna).border.top.style == "medium"
    for fila in range(PRIMERA_FILA_DATOS, PRIMERA_FILA_DATOS + len(registros)):
        for columna in range(1, columnas + 1):
            borde = hoja.cell(row=fila, column=columna).border
            assert borde.top.style == "thin" and borde.left.style == "thin"
    # Los dos banners combinados de la fila 1 (D–H e I–L) también van bordeados.
    for columna in range(4, 13):
        assert hoja.cell(row=1, column=columna).border.top.style == "medium"
